import os
import openpyxl

# Jake Pan

def get_cell(input_dir):
    try:
        for root, dirs, files in os.walk(input_dir):
            sum_t1 = 0
            sum_t2 = 0
            count = 0
            # dic1 = {}
            # data_list_total = []
            for file in files:
                    file_dir_name = os.path.join(root, file)
                    print(file_dir_name)
                    wb = openpyxl.load_workbook(file_dir_name, data_only=True)
                    # data_only for not return formula (ex: sum[A1:G1])

                    sheet = wb.active
                    num_row = sheet.max_row
                    num_column = sheet.max_column

                    target_number1 = sheet['G27'].value  # 支数
                    target_number2 = sheet['I27'].value  # 重量
                    file_name_str = str(file)
                    data_list1 = [target_number1, target_number2]

                    print("列表行数：" + str(num_row))
                    print("列表列数：" + str(num_column))
                    print("数量：" + str(target_number1))
                    print("总计：" + str(target_number2))
                    print("文件名：" + file_name_str)
                    print("数据集：" + str(data_list1))

                    sum_t1 = sum_t1 + target_number1
                    sum_t2 = sum_t2 + target_number2
                    count = count + 1
                    #dict or list .append
                    # return data_list_total

                    # serial_number = sheet['A4':'A27']  # 序列号
                    # print("序列号：" + str(serial_number))

        print("\n")
        print("===============================================")
        print("===============================================")
        print("\n")
        print("文件处理数：" + str(count))
        print("月核对数量：" + str(sum_t1))
        print("月核对总计：" + str(sum_t2))
        return 123
        # print("月核对合集：" + str(data_list_total))

    except Exception as er1:
        print("获得名字出错 [E1]，错误为： " + er1)



# 程序入口
if __name__ == "__main__":
    # 设置导入和输出的文件夹
    # 填入路径 absolute path
    input_filePath = "C:\\Users\\Jiacong96\\Desktop\\scratch_s\\python_exercise\\workStats\\M7M8M9\\"
    input_filePath1 = "C:\\general\\script\\workStats\\M7M8M9\\"
    output_filePath = "C:\\Users\\Jiacong96\\Desktop\\scratch_s\\python_exercise\\workStats\\checkList_month\\"
    # file_name_list = get_file_name(input_filePath)

    cell_value_list = get_cell(input_filePath1)
    print(cell_value_list)





