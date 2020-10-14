import openpyxl
wb = openpyxl.load_workbook(filePath)
sheet_names = wb.get_sheet_names()
sheet = wb.get_sheet_by_name(sheet_names[0]) #获取第一张表
#此方法会报错，用最新 sheet= wb[“sheet”]
numRows = sheet.max_row
numCols = sheet.max_column
for m in range(1, numRows+1):
    for n in range(1, numCols+1):
        print(sheet.cell(row = m, column = n).vlaue) # sheet.cell.value 获取数值
# 已经知道单元格的坐标，查询坐标的数值
# 此方法会查询excel返回初始的公式
